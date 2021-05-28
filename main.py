from argparse import ArgumentParser

from browser import Browser
from time import sleep
from openpyxl import load_workbook

import sys
from os import path, getenv
from dotenv import load_dotenv

from datetime import datetime

PROGRAM = "Это должен знать каждый"
MODULE = "Это должен"


def main(args):
    sys.stdout = open('latest.log', 'w')
    browser = Browser()
    wordbook = args.wordbook
    _type = args.type

    # Авторизация на сайте
    try:
        browser.auth(getenv("LOGIN"), getenv("PASSWORD"))
        print("Авторизация успешна!")
        sleep(1)
    except:
        print("Неудалось авторизоваться! Проверьте правильность Логина и Пароля!")
        exit(1)

    if _type == 0:
        add_contract(browser, wordbook)
    elif _type == 1:
        check_groups(browser, wordbook)

    browser.close()
    sys.stdout.close()


def add_contract(browser, wordbook):
    browser.add_contract_setup(
        program=PROGRAM, module=MODULE, contract_date=datetime(2021, 6, 1))
    for row in wordbook.active.iter_rows(min_row=2, max_col=5, values_only=True):
        if row[0] is None:
            continue
        try:
            browser.add_contract(str(row[0]), row[2], row[3], row[4], row[1])
        except Exception as e:
            print(e)
            print("Перехожу к следующему")


def check_groups(browser, wordbook):
    for ws_name in wordbook.sheetnames:
        ws = wordbook[ws_name]
        print("Группа: %s \n" % ws_name)

        # Получение данных с сайта
        try:
            group_name, group_members, group_members_count = browser.get_group_info(
                ws['B1'].value)
        except Exception as e:
            print("Произошла ошибка\n")
            print(e)
            continue

        # Проверка на разные группы
        if group_name != ws_name:
            print("Запрос по группе %s класса, вернул %s класс" %
                  (ws_name, group_name))
            continue

        members = list()
        # Подгрузка всех учеников из листа
        for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0] is None:
                continue
            members.append(row[0])

        # Сортировка найденых и не найденных обучающихся
        not_found = list(
            set(sorted(members, key=lambda x: ord(x[0]))) - set(sorted(group_members, key=lambda x: ord(x[0]))))

        if len(not_found) > 1:
            print("X | В группе: (в листе: %d) (на сайте:%d) | не найдены следующие обучающиеся:" %
                  (len(members), group_members_count))
            print("\n".join(not_found))
        else:
            print("+ | В группе: (в листе: %d) (на сайте:%d) | Проверка успешно пройдена!" %
                  (len(members), group_members_count))


def load_env(file_name=".env"):
    dotenv_path = path.join(path.dirname(__file__), file_name)
    if path.exists(dotenv_path):
        load_dotenv(dotenv_path)


def arg_parser():
    parser = ArgumentParser(description="Данная программа предназначена для более удобного и быстрого добавления "
                                        "списка ПФДО обучающихся в Специализрованные секции")

    parser.add_argument("-wb", "--wordbook", metavar="список.xlsx", required=True, type=load_workbook,
                        help="Файл с данными (Excel) [Пример файла в папке example]")
    parser.add_argument("-t", "--type", type=int, metavar="0/1", required=True, choices=[0, 1],
                        help="Тип работы программы (0 - добавление/1 - сверка)")
    return parser


if __name__ == "__main__":
    load_env()
    parser = arg_parser()
    main(parser.parse_args())
